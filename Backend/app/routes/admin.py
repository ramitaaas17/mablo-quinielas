import csv
import io
import os
import re
import uuid
import base64
import datetime
import logging
from functools import wraps
from flask import Blueprint, request, jsonify, Response, send_from_directory
from flask_jwt_extended import jwt_required, get_jwt_identity

audit_logger = logging.getLogger("audit")
from sqlalchemy import text
from werkzeug.security import generate_password_hash
from ..models import (
    Quiniela, Partido, UsuarioQuiniela, Prediccion,
    Usuario, Pago, Liga, Equipo
)
from ..extensions import db, limiter

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)
MAX_IMAGE_BYTES = 15 * 1024 * 1024  # 15 MB máx
ALLOWED_TYPES = {'image/jpeg', 'image/png', 'image/webp', 'image/gif'}


def _save_imagen(data_url: str) -> str:
    """Recibe un data-URL Base64, guarda el archivo en disco y devuelve la URL relativa."""
    if not data_url or not data_url.startswith('data:'):
        return data_url  # Ya es una ruta, devuélvela tal cual
    try:
        header, encoded = data_url.split(',', 1)
        # Validar tipo de imagen
        mime = header.split(';')[0].replace('data:', '')
        if mime not in ALLOWED_TYPES:
            raise ValueError(f"Tipo de imagen no permitido: {mime}")
        raw = base64.b64decode(encoded)
        # Validar tamaño
        if len(raw) > MAX_IMAGE_BYTES:
            raise ValueError("La imagen supera el límite de 5 MB")
        ext = 'jpg'
        if 'png' in header:
            ext = 'png'
        elif 'webp' in header:
            ext = 'webp'
        elif 'gif' in header:
            ext = 'gif'
        filename = f"{uuid.uuid4().hex}.{ext}"
        filepath = os.path.join(UPLOAD_DIR, filename)
        with open(filepath, 'wb') as f:
            f.write(raw)
        return f"/uploads/{filename}"
    except Exception:
        return data_url  # Si falla, guarda lo que tenga

admin_bp = Blueprint('admin', __name__)


# ─── DECORATOR ────────────────────────────────────────────────────────────────

def admin_required(fn):
    @wraps(fn)
    @jwt_required()
    def wrapper(*args, **kwargs):
        id_usr = get_jwt_identity()
        user = Usuario.query.get(id_usr)
        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404
        if not user.is_admin:
            return jsonify({"error": "Acceso denegado. Se requieren permisos de administrador."}), 403
        return fn(*args, **kwargs)
    return wrapper


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _initials(nombre):
    partes = nombre.split()
    return ''.join(x[0].upper() for x in partes[:2] if x)


def _format_quiniela(q):
    liga = Liga.query.get(q.id_liga)
    num_partidos = Partido.query.filter_by(id_quiniela=q.id_quiniela).count()
    num_jugadores = UsuarioQuiniela.query.filter_by(id_quiniela=q.id_quiniela).count()
    pagos_confirmados = Pago.query.filter_by(id_quiniela=q.id_quiniela, estado='confirmado').count()
    pagos_pendientes = Pago.query.filter_by(id_quiniela=q.id_quiniela, estado='pendiente').count()
    return {
        "id": str(q.id_quiniela),
        "nombre": q.nombre,
        "liga_nombre": liga.nombre if liga else "",
        "inicio": q.inicio.isoformat(),
        "cierre": q.cierre.isoformat(),
        "precio_entrada": float(q.precio_entrada),
        "comision": float(q.comision),
        "pozo_acumulado": float(q.pozo_acumulado),
        "estado": q.estado,
        "imagen_fondo": q.imagen_fondo,
        "num_partidos": num_partidos,
        "num_jugadores": num_jugadores,
        "pagos_confirmados": pagos_confirmados,
        "pagos_pendientes": pagos_pendientes,
    }


def _format_partido_admin(p):
    eq_local = Equipo.query.get(p.local)
    eq_visit = Equipo.query.get(p.visitante)
    return {
        "id": str(p.id_partido),
        "local_id": str(p.local),
        "local_nombre": eq_local.nombre if eq_local else "",
        "visitante_id": str(p.visitante),
        "visitante_nombre": eq_visit.nombre if eq_visit else "",
        "inicio": p.inicio.isoformat(),
        "ptos_local": p.ptos_local,
        "ptos_visitante": p.ptos_visitante,
        "cancelado": p.cancelado,
    }


def _format_pago(pg):
    usr = Usuario.query.get(pg.id_usr)
    confirmador = Usuario.query.get(pg.confirmado_por) if pg.confirmado_por else None
    return {
        "id_pago": str(pg.id_pago),
        "id_usr": str(pg.id_usr),
        "nombre": usr.nombre_completo if usr else "",
        "correo": usr.correo if usr else "",
        "initials": _initials(usr.nombre_completo) if usr else "??",
        "foto_perfil": usr.foto_perfil if usr else None,
        "monto": float(pg.monto),
        "metodo": pg.metodo,
        "estado": pg.estado,
        "nota": pg.nota,
        "fecha_pago": pg.fecha_pago.isoformat(),
        "fecha_confirmacion": pg.fecha_confirmacion.isoformat() if pg.fecha_confirmacion else None,
        "confirmado_por_nombre": confirmador.nombre_completo if confirmador else None,
    }


# ─── STATS ────────────────────────────────────────────────────────────────────

@admin_bp.route('/stats', methods=['GET'])
@admin_required
def get_stats():
    total_usuarios = Usuario.query.filter_by(is_admin=False).count()
    quinielas_activas = Quiniela.query.filter_by(estado='abierta').count()
    pagos_confirmados = Pago.query.filter_by(estado='confirmado').count()
    pagos_pendientes = Pago.query.filter_by(estado='pendiente').count()

    # Pozo total activo
    qs_activas = Quiniela.query.filter_by(estado='abierta').all()
    pozo_total = sum(float(q.pozo_acumulado) for q in qs_activas)

    return jsonify({
        "total_usuarios": total_usuarios,
        "quinielas_activas": quinielas_activas,
        "pagos_confirmados": pagos_confirmados,
        "pagos_pendientes": pagos_pendientes,
        "pozo_total": pozo_total,
    }), 200


# ─── ACTIVIDAD RECIENTE ───────────────────────────────────────────────────────

@admin_bp.route('/actividad', methods=['GET'])
@admin_required
def get_actividad():
    # Predicciones: agrupar por (user, quiniela) para no mostrar una fila por predicción
    preds_rows = (
        db.session.query(Prediccion.id_usr, Prediccion.fecha, Partido.id_quiniela)
        .join(Partido, Prediccion.id_partido == Partido.id_partido)
        .order_by(Prediccion.fecha.desc())
        .limit(500)
        .all()
    )

    # Agrupar: una fila por (usuario, quiniela)
    grupos = {}  # (user_id, quiniela_id) -> { count, latest }
    for id_usr, fecha, id_quiniela in preds_rows:
        key = (str(id_usr), str(id_quiniela))
        if key not in grupos:
            grupos[key] = {"count": 0, "latest": fecha}
        grupos[key]["count"] += 1
        if fecha > grupos[key]["latest"]:
            grupos[key]["latest"] = fecha

    actividad = []
    for (user_id, quiniela_id), v in grupos.items():
        usr = Usuario.query.get(user_id)
        quiniela = Quiniela.query.get(quiniela_id)
        if usr and quiniela:
            actividad.append({
                "tipo": "Predicción",
                "usuario": usr.nombre_completo,
                "initials": _initials(usr.nombre_completo),
                "foto_perfil": usr.foto_perfil or None,
                "quiniela": quiniela.nombre,
                "fecha": v["latest"].isoformat(),
                "count": v["count"],
            })

    # Pagos
    pagos = Pago.query.order_by(Pago.fecha_pago.desc()).limit(20).all()
    for pg in pagos:
        usr = Usuario.query.get(pg.id_usr)
        quiniela = Quiniela.query.get(pg.id_quiniela)
        if usr:
            actividad.append({
                "tipo": "Pago enviado" if pg.estado == 'pendiente' else "Pago confirmado",
                "usuario": usr.nombre_completo,
                "initials": _initials(usr.nombre_completo),
                "foto_perfil": usr.foto_perfil or None,
                "quiniela": quiniela.nombre if quiniela else "",
                "fecha": pg.fecha_pago.isoformat(),
                "count": 1,
            })

    actividad.sort(key=lambda x: x['fecha'], reverse=True)
    return jsonify(actividad[:15]), 200


# ─── QUINIELAS ────────────────────────────────────────────────────────────────

@admin_bp.route('/quinielas', methods=['GET'])
@admin_required
def get_quinielas_admin():
    quinielas = Quiniela.query.order_by(Quiniela.inicio.desc()).all()
    return jsonify([_format_quiniela(q) for q in quinielas]), 200


@admin_bp.route('/quinielas/<id_quiniela>/codigo-invitacion', methods=['GET'])
@admin_required
def get_codigo_invitacion(id_quiniela):
    """Obtiene o genera el código de invitación único para la quiniela."""
    import random, string
    from ..models import Invitacion
    Quiniela.query.get_or_404(id_quiniela)
    inv = Invitacion.query.filter_by(id_quiniela=id_quiniela).first()
    if not inv:
        # Generar código único de 8 chars alfanumérico mayúscula
        while True:
            codigo = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            if not Invitacion.query.filter_by(codigo=codigo).first():
                break
        inv = Invitacion(id_quiniela=id_quiniela, codigo=codigo)
        db.session.add(inv)
        db.session.commit()

    return jsonify({
        "codigo": inv.codigo,
        "activo": inv.activo,
        "url": f"/unirse/{inv.codigo}",
    }), 200


@admin_bp.route('/quinielas', methods=['POST'])
@admin_required
def crear_quiniela():
    data = request.get_json(silent=True) or {}
    campos = ['nombre', 'id_liga', 'cierre', 'precio_entrada', 'comision']
    for campo in campos:
        if data.get(campo) is None:
            return jsonify({"error": f"Falta el campo: {campo}"}), 400

    try:
        q = Quiniela(
            nombre=data['nombre'],
            id_liga=data['id_liga'],
            inicio=datetime.datetime.now(),
            cierre=datetime.datetime.fromisoformat(data['cierre']).astimezone(datetime.timezone.utc).replace(tzinfo=None),
            precio_entrada=float(data['precio_entrada']),
            comision=float(data['comision']),
            estado='abierta',
            imagen_fondo=_save_imagen(data.get('imagen_fondo')),
        )
        db.session.add(q)
        db.session.commit()
        return jsonify({"mensaje": "Quiniela creada", "id": str(q.id_quiniela)}), 201
    except ValueError as e:
        return jsonify({"error": f"Datos inválidos: {str(e)}"}), 400
    except Exception as e:
        db.session.rollback()
        import traceback
        logging.getLogger(__name__).error("Error al crear quiniela: %s\n%s", e, traceback.format_exc())
        return jsonify({"error": "Error al crear quiniela"}), 500


@admin_bp.route('/quinielas/<id_quiniela>', methods=['GET'])
@admin_required
def get_quiniela_admin(id_quiniela):
    q = Quiniela.query.get_or_404(id_quiniela)
    partidos = Partido.query.filter_by(id_quiniela=id_quiniela).order_by(Partido.inicio.asc()).all()
    resultado = _format_quiniela(q)
    resultado['partidos'] = [_format_partido_admin(p) for p in partidos]
    return jsonify(resultado), 200


@admin_bp.route('/quinielas/<id_quiniela>', methods=['PATCH'])
@admin_required
def editar_quiniela(id_quiniela):
    q = Quiniela.query.get_or_404(id_quiniela)
    if q.estado == 'resuelta':
        return jsonify({"error": "No se puede editar una quiniela resuelta"}), 400

    data = request.get_json(silent=True) or {}
    if 'nombre' in data:
        q.nombre = data['nombre']
    if 'cierre' in data:
        q.cierre = datetime.datetime.fromisoformat(data['cierre'])
    if 'precio_entrada' in data:
        nuevo_precio = float(data['precio_entrada'])
        if q.precio_entrada != nuevo_precio:
            q.precio_entrada = nuevo_precio
            # Actualizar pagos pendientes al nuevo precio
            pagos_pendientes = Pago.query.filter_by(id_quiniela=id_quiniela, estado='pendiente').all()
            for p in pagos_pendientes:
                p.monto = nuevo_precio
    if 'comision' in data:
        q.comision = float(data['comision'])
    if 'imagen_fondo' in data:
        q.imagen_fondo = _save_imagen(data['imagen_fondo'])

    db.session.commit()
    return jsonify({"mensaje": "Quiniela actualizada"}), 200


@admin_bp.route('/quinielas/<id_quiniela>/cerrar', methods=['PATCH'])
@admin_required
def cerrar_quiniela(id_quiniela):
    q = Quiniela.query.get_or_404(id_quiniela)
    if q.estado != 'abierta':
        return jsonify({"error": "La quiniela no está abierta"}), 400
    q.estado = 'cerrada'
    db.session.commit()
    return jsonify({"mensaje": "Quiniela cerrada"}), 200


@admin_bp.route('/quinielas/<id_quiniela>', methods=['DELETE'])
@admin_required
def eliminar_quiniela(id_quiniela):
    q = Quiniela.query.get_or_404(id_quiniela)
    try:
        db.session.delete(q)
        db.session.commit()
        audit_logger.info("admin elimino quiniela %s (%s)", id_quiniela, q.nombre)
        return jsonify({"mensaje": "Quiniela eliminada definitivamente"}), 200
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Error al eliminar quiniela"}), 500


@admin_bp.route('/quinielas/<id_quiniela>/partidos', methods=['POST'])
@admin_required
def agregar_partidos(id_quiniela):
    q = Quiniela.query.get_or_404(id_quiniela)
    if q.estado == 'resuelta':
        return jsonify({"error": "No se pueden agregar partidos a una quiniela resuelta"}), 400

    data = request.get_json(silent=True) or {}
    partidos_data = data.get('partidos', [])
    if not partidos_data:
        return jsonify({"error": "Se requiere al menos un partido"}), 400

    creados = []
    try:
        for pd in partidos_data:
            if not pd.get('local') or not pd.get('visitante') or not pd.get('inicio'):
                return jsonify({"error": "Cada partido necesita local, visitante e inicio"}), 400
            if pd['local'] == pd['visitante']:
                return jsonify({"error": "Local y visitante deben ser equipos distintos"}), 400

            partido = Partido(
                id_quiniela=id_quiniela,
                local=pd['local'],
                visitante=pd['visitante'],
                inicio=datetime.datetime.fromisoformat(pd['inicio']),
            )
            db.session.add(partido)
            creados.append(partido)

        db.session.commit()
        return jsonify({
            "mensaje": f"{len(creados)} partido(s) agregado(s)",
            "ids": [str(p.id_partido) for p in creados]
        }), 201
    except ValueError as e:
        db.session.rollback()
        return jsonify({"error": f"Fecha inválida: {str(e)}"}), 400
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Error al agregar partidos"}), 500


@admin_bp.route('/quinielas/<id_quiniela>/participantes', methods=['GET'])
@admin_required
def obtener_participantes(id_quiniela):
    Quiniela.query.get_or_404(id_quiniela)
    participantes = db.session.query(UsuarioQuiniela, Usuario).join(
        Usuario, UsuarioQuiniela.id_usr == Usuario.id_usr
    ).filter(UsuarioQuiniela.id_quiniela == id_quiniela).all()

    resultado = []
    for uq, usr in participantes:
        pago = Pago.query.filter_by(id_usr=usr.id_usr, id_quiniela=id_quiniela).first()
        resultado.append({
            "id": str(usr.id_usr),
            "nombre": usr.nombre_completo,
            "correo": usr.correo,
            "initials": _initials(usr.nombre_completo),
            "foto_perfil": usr.foto_perfil or None,
            "puntos": uq.puntos_total,
            "estado_pago": pago.estado if pago else "sin_pago",
            "metodo_pago": pago.metodo if pago else None,
            "fecha_union": uq.fecha_union.isoformat(),
        })
    return jsonify(resultado), 200


@admin_bp.route('/quinielas/<id_quiniela>/predicciones', methods=['GET'])
@admin_required
def ver_predicciones(id_quiniela):
    Quiniela.query.get_or_404(id_quiniela)
    partidos = Partido.query.filter_by(id_quiniela=id_quiniela).order_by(Partido.inicio.asc()).all()
    participantes = UsuarioQuiniela.query.filter_by(id_quiniela=id_quiniela).all()

    resultado = []
    for uq in participantes:
        usr = Usuario.query.get(uq.id_usr)
        if not usr:
            continue
        preds_usr = []
        for partido in partidos:
            pred = Prediccion.query.filter_by(id_usr=uq.id_usr, id_partido=partido.id_partido).first()
            preds_usr.append({
                "id_partido": str(partido.id_partido),
                "selecciones": pred.selecciones if pred else [],
                "es_correcta": pred.es_correcta if pred else None,
            })
        resultado.append({
            "id_usr": str(usr.id_usr),
            "nombre": usr.nombre_completo,
            "initials": _initials(usr.nombre_completo),
            "puntos": uq.puntos_total,
            "predicciones": preds_usr,
        })
    return jsonify(resultado), 200


@admin_bp.route('/quinielas/<id_quiniela>/resolver', methods=['POST'])
@admin_required
def resolver_quiniela(id_quiniela):
    q = Quiniela.query.get_or_404(id_quiniela)
    if q.estado == 'resuelta':
        return jsonify({"error": "Ya fue resuelta"}), 400

    partidos = Partido.query.filter_by(id_quiniela=id_quiniela, cancelado=False).all()
    for p in partidos:
        if p.ptos_local is None or p.ptos_visitante is None:
            return jsonify({"error": f"El partido {p.id_partido} no tiene resultado todavía"}), 400

    try:
        query = text("SELECT * FROM resolver_quiniela(:id_q)")
        rows = db.session.execute(query, {"id_q": id_quiniela}).fetchall()
        db.session.commit()

        db.session.refresh(q)

        # Procesar resultados de la función SQL
        resultados = []
        ganador = None
        max_pts = 0

        for row in rows:
            pts = row[2]
            es_gan = row[3]
            if pts > max_pts:
                max_pts = pts
            if es_gan:
                ganador = {"nombre": row[1], "puntos": pts}
            resultados.append({
                "id_usr": str(row[0]),
                "nombre": row[1],
                "puntos": pts,
                "es_ganador": es_gan,
            })

        # Determinar si hay empate (2+ con max puntos)
        con_max = [r for r in resultados if r['puntos'] == max_pts]
        hay_empate = len(con_max) > 1

        premio = 0.0
        if ganador and not hay_empate:
            premio = float(q.pozo_acumulado) * (1 - float(q.comision) / 100)

        return jsonify({
            "mensaje": "Quiniela resuelta",
            "hay_empate": hay_empate,
            "ganador": ganador,
            "premio": premio,
            "pozo_acumulado": float(q.pozo_acumulado),
            "resultados": resultados,
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


# ─── PAGOS ────────────────────────────────────────────────────────────────────

@admin_bp.route('/quinielas/<id_quiniela>/pagos', methods=['GET'])
@admin_required
def listar_pagos(id_quiniela):
    Quiniela.query.get_or_404(id_quiniela)
    pagos = Pago.query.filter_by(id_quiniela=id_quiniela).all()
    return jsonify([_format_pago(p) for p in pagos]), 200


@admin_bp.route('/pagos/<id_pago>/confirmar', methods=['PATCH'])
@admin_required
def confirmar_pago(id_pago):
    id_admin = get_jwt_identity()
    pago = Pago.query.get_or_404(id_pago)
    if pago.estado == 'confirmado':
        return jsonify({"error": "El pago ya fue confirmado"}), 400

    data = request.get_json(silent=True) or {}
    pago.estado = 'confirmado'
    pago.metodo = data.get('metodo', 'efectivo')
    pago.nota = data.get('nota')
    pago.monto = float(data.get('monto', pago.monto))
    pago.fecha_confirmacion = datetime.datetime.utcnow()
    pago.confirmado_por = id_admin

    # Actualizar monto pagado en usuario_quiniela
    uq = UsuarioQuiniela.query.filter_by(
        id_usr=pago.id_usr,
        id_quiniela=pago.id_quiniela
    ).first()
    if uq:
        uq.monto_pagado = pago.monto

    # Actualizar pozo acumulado
    q = Quiniela.query.get(pago.id_quiniela)
    if q:
        total_confirmado = db.session.query(
            db.func.sum(Pago.monto)
        ).filter_by(id_quiniela=pago.id_quiniela, estado='confirmado').scalar() or 0
        q.pozo_acumulado = float(total_confirmado)

    db.session.commit()
    audit_logger.info("[AUDIT] CONFIRMAR_PAGO admin=%s pago=%s usuario=%s monto=%s quiniela=%s",
                      id_admin, id_pago, pago.id_usr, pago.monto, pago.id_quiniela)
    return jsonify({"mensaje": "Pago confirmado", "pago": _format_pago(pago)}), 200


@admin_bp.route('/pagos/<id_pago>/rechazar', methods=['PATCH'])
@admin_required
def rechazar_pago(id_pago):
    id_admin = get_jwt_identity()
    pago = Pago.query.get_or_404(id_pago)
    if pago.estado == 'rechazado':
        return jsonify({"error": "El pago ya fue rechazado"}), 400

    data = request.get_json(silent=True) or {}
    pago.estado = 'rechazado'
    pago.nota = data.get('nota')
    pago.fecha_confirmacion = datetime.datetime.utcnow()
    pago.confirmado_por = id_admin

    db.session.commit()
    return jsonify({"mensaje": "Pago rechazado"}), 200


@admin_bp.route('/pagos/<id_pago>/revertir', methods=['PATCH'])
@admin_required
def revertir_pago(id_pago):
    pago = Pago.query.get_or_404(id_pago)
    pago.estado = 'pendiente'
    pago.fecha_confirmacion = None
    pago.confirmado_por = None
    pago.nota = None

    uq = UsuarioQuiniela.query.filter_by(
        id_usr=pago.id_usr,
        id_quiniela=pago.id_quiniela
    ).first()
    if uq:
        uq.monto_pagado = 0

    q = Quiniela.query.get(pago.id_quiniela)
    if q:
        total_confirmado = db.session.query(
            db.func.sum(Pago.monto)
        ).filter_by(id_quiniela=pago.id_quiniela, estado='confirmado').scalar() or 0
        q.pozo_acumulado = float(total_confirmado)

    db.session.commit()
    return jsonify({"mensaje": "Pago revertido a pendiente"}), 200


# ─── PARTIDOS ─────────────────────────────────────────────────────────────────

@admin_bp.route('/partidos/proximos', methods=['GET'])
@admin_required
def get_proximos_partidos():
    """Retorna próximos partidos de Liga MX desde SofaScore (para importar)."""
    try:
        from ..scraper import get_proximos
        todos = get_proximos()
        proximos = [p for p in todos if p['estado_tipo'] == 'notstarted']
        return jsonify(proximos[:30]), 200
    except Exception as e:
        return jsonify({"error": str(e), "partidos": []}), 500


@admin_bp.route('/quinielas/<id_quiniela>/importar-partidos', methods=['POST'])
@admin_required
def importar_partidos(id_quiniela):
    """
    Importa partidos del scraper a la quiniela.
    Payload: { partidos: [{ local, visitante, fecha_iso }] }
    Hace matching automático de nombres con equipos en DB.
    Si un equipo no existe, lo crea en la liga de la quiniela.
    """
    import unicodedata as _uc

    q = Quiniela.query.get_or_404(id_quiniela)
    if q.estado == 'resuelta':
        return jsonify({"error": "No se pueden agregar partidos a una quiniela resuelta"}), 400

    data = request.get_json(silent=True) or {}
    partidos_data = data.get('partidos', [])
    if not partidos_data:
        return jsonify({"error": "Se requiere al menos un partido"}), 400

    from ..scheduler import match_nombre

    equipos_liga = Equipo.query.filter_by(id_liga=q.id_liga).all()

    def find_or_create(nombre_scraper: str) -> Equipo:
        for eq in equipos_liga:
            if match_nombre(nombre_scraper, eq.nombre):
                return eq
        # No existe: crear con nombre original del scraper
        nuevo = Equipo(id_liga=q.id_liga, nombre=nombre_scraper)
        db.session.add(nuevo)
        db.session.flush()
        equipos_liga.append(nuevo)
        return nuevo

    creados = []
    errores = []
    try:
        for pd in partidos_data:
            local_nom  = pd.get('local', '').strip()
            visita_nom = pd.get('visitante', '').strip()
            fecha_iso  = pd.get('fecha_iso', '')

            if not local_nom or not visita_nom or not fecha_iso:
                errores.append(f"Partido incompleto: {pd}")
                continue

            eq_local = find_or_create(local_nom)
            eq_visit = find_or_create(visita_nom)

            if eq_local.id_eqp == eq_visit.id_eqp:
                errores.append(f"Local y visitante son el mismo equipo: {local_nom}")
                continue

            # Evitar duplicados (mismo local+visitante+fecha en esta quiniela)
            existe = Partido.query.filter_by(
                id_quiniela=id_quiniela,
                local=eq_local.id_eqp,
                visitante=eq_visit.id_eqp,
            ).first()
            if existe:
                continue

            partido = Partido(
                id_quiniela=id_quiniela,
                local=eq_local.id_eqp,
                visitante=eq_visit.id_eqp,
                inicio=datetime.datetime.fromisoformat(fecha_iso),
            )
            db.session.add(partido)
            creados.append(partido)

        db.session.commit()
        return jsonify({
            "mensaje": f"{len(creados)} partido(s) importado(s)",
            "ids": [str(p.id_partido) for p in creados],
            "errores": errores,
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@admin_bp.route('/partidos/<id_partido>/marcador', methods=['PATCH'])
@admin_required
def actualizar_marcador(id_partido):
    """Permite al admin ingresar o corregir el marcador de un partido manualmente."""
    partido = Partido.query.get_or_404(id_partido)
    q = Quiniela.query.get(partido.id_quiniela)
    if q and q.estado == 'resuelta':
        return jsonify({"error": "La quiniela ya fue resuelta"}), 400
    if partido.cancelado:
        return jsonify({"error": "El partido está cancelado"}), 400

    data = request.get_json(silent=True) or {}
    g_local   = data.get('goles_local')
    g_visita  = data.get('goles_visitante')

    if g_local is None or g_visita is None:
        return jsonify({"error": "Se requieren goles_local y goles_visitante"}), 400
    try:
        g_local  = int(g_local)
        g_visita = int(g_visita)
        if g_local < 0 or g_visita < 0:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"error": "Los goles deben ser números enteros no negativos"}), 400

    partido.ptos_local     = g_local
    partido.ptos_visitante = g_visita
    db.session.commit()
    return jsonify({"mensaje": "Marcador actualizado", "partido": _format_partido_admin(partido)}), 200


@admin_bp.route('/partidos/<id_partido>/cancelar', methods=['PATCH'])
@admin_required
def cancelar_partido(id_partido):
    partido = Partido.query.get_or_404(id_partido)
    q = Quiniela.query.get(partido.id_quiniela)
    if q and q.estado == 'resuelta':
        return jsonify({"error": "La quiniela ya fue resuelta"}), 400

    partido.cancelado = True
    db.session.commit()
    return jsonify({"mensaje": "Partido cancelado"}), 200


@admin_bp.route('/partidos/<id_partido>/descancelar', methods=['PATCH'])
@admin_required
def descancelar_partido(id_partido):
    partido = Partido.query.get_or_404(id_partido)
    q = Quiniela.query.get(partido.id_quiniela)
    if q and q.estado == 'resuelta':
        return jsonify({"error": "La quiniela ya fue resuelta"}), 400
    if not partido.cancelado:
        return jsonify({"error": "El partido no está cancelado"}), 400

    partido.cancelado = False
    db.session.commit()
    return jsonify({"mensaje": "Partido reactivado", "partido": _format_partido_admin(partido)}), 200


# ─── USUARIOS ─────────────────────────────────────────────────────────────────

@admin_bp.route('/usuarios', methods=['GET'])
@admin_required
def listar_usuarios():
    usuarios = Usuario.query.filter_by(is_admin=False).order_by(
        Usuario.fecha_creacion.desc()
    ).all()
    resultado = []
    for u in usuarios:
        num_quinielas = UsuarioQuiniela.query.filter_by(id_usr=u.id_usr).count()
        total_puntos = db.session.query(
            db.func.sum(UsuarioQuiniela.puntos_total)
        ).filter_by(id_usr=u.id_usr).scalar() or 0
        eq = Equipo.query.get(u.equipo_favorito) if u.equipo_favorito else None
        resultado.append({
            "id": str(u.id_usr),
            "nombre": u.nombre_completo,
            "username": u.username,
            "correo": u.correo,
            "initials": _initials(u.nombre_completo),
            "foto_perfil": u.foto_perfil or None,
            "equipo_favorito": eq.nombre if eq else None,
            "num_quinielas": num_quinielas,
            "total_puntos": int(total_puntos),
            "fecha_creacion": u.fecha_creacion.isoformat(),
            "is_admin": u.is_admin,
        })
    return jsonify(resultado), 200


_ADMIN_EMAIL_RE    = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
_ADMIN_PASSWORD_RE = re.compile(r'^(?=.*[A-Za-z])(?=.*\d).{8,}$')
_ADMIN_USERNAME_RE = re.compile(r'^[a-zA-Z0-9_]{3,50}$')

@admin_bp.route('/usuarios', methods=['POST'])
@admin_required
def crear_usuario():
    data = request.get_json(silent=True) or {}
    campos = ['nombre_completo', 'username', 'correo', 'contrasena', 'fecha_nacimiento']
    for campo in campos:
        if not data.get(campo):
            return jsonify({"error": f"Falta el campo: {campo}"}), 400

    if not _ADMIN_EMAIL_RE.match(data['correo']):
        return jsonify({"error": "Correo inválido"}), 400

    if not _ADMIN_USERNAME_RE.match(data['username']):
        return jsonify({"error": "Username solo puede contener letras, números y guión bajo (3–50 caracteres)"}), 400

    if not _ADMIN_PASSWORD_RE.match(data['contrasena']):
        return jsonify({"error": "La contraseña debe tener al menos 8 caracteres, incluir letras y números"}), 400

    if Usuario.query.filter_by(username=data['username']).first():
        return jsonify({"error": "Username ya en uso"}), 409
    if Usuario.query.filter_by(correo=data['correo']).first():
        return jsonify({"error": "Correo ya en uso"}), 409

    try:
        nuevo = Usuario(
            nombre_completo=data['nombre_completo'],
            username=data['username'],
            correo=data['correo'],
            contraseña_hasheada=generate_password_hash(data['contrasena']),
            fecha_nacimiento=datetime.date.fromisoformat(data['fecha_nacimiento']),
            equipo_favorito=data.get('equipo_favorito') or None,
            is_admin=bool(data.get('is_admin', False)),
        )
        db.session.add(nuevo)
        db.session.commit()
        audit_logger.info("[AUDIT] CREAR_USUARIO admin=%s creó usuario=%s (%s) is_admin=%s",
                          get_jwt_identity(), str(nuevo.id_usr), nuevo.correo, nuevo.is_admin)
        return jsonify({"mensaje": "Usuario creado", "id": str(nuevo.id_usr)}), 201
    except ValueError:
        return jsonify({"error": "Fecha de nacimiento inválida"}), 400
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Error al crear usuario"}), 500


@admin_bp.route('/usuarios/<id_usuario>', methods=['DELETE'])
@admin_required
def eliminar_usuario(id_usuario):
    admin_id = get_jwt_identity()
    user = Usuario.query.get_or_404(id_usuario)
    if user.is_admin:
        return jsonify({"error": "No se puede eliminar un administrador"}), 400
    audit_logger.warning("[AUDIT] ELIMINAR_USUARIO admin=%s eliminó usuario=%s (%s)", admin_id, id_usuario, user.correo)
    db.session.delete(user)
    db.session.commit()
    return jsonify({"mensaje": "Usuario eliminado"}), 200


# ─── PALETA DE DISEÑO ─────────────────────────────────────────────────────────

_C = {
    'dark':        '#1a1a1a',
    'green':       '#3dbb78',
    'green_dark':  '#25854f',
    'green_light': '#d6f5e8',
    'orange':      '#f4a030',
    'orange_light':'#fff3e0',
    'red':         '#ef4444',
    'red_light':   '#fee2e2',
    'gray':        '#6b6b6b',
    'border':      '#e4e4e0',
    'bg':          '#fafaf8',
    'bg2':         '#f2f2ef',
    'white':       '#ffffff',
    'gold':        '#d97706',
    'gold_light':  '#fef3c7',
    'silver':      '#64748b',
    'silver_light':'#f1f5f9',
    'bronze':      '#92400e',
    'bronze_light':'#fde8ca',
}


# ─── EXCEL HELPERS ────────────────────────────────────────────────────────────

def _xl_font(bold=False, size=10, color='1A1A1A', name='Calibri'):
    from openpyxl.styles import Font
    return Font(name=name, bold=bold, size=size, color=color.lstrip('#').upper())

def _xl_fill(hex_color):
    from openpyxl.styles import PatternFill
    return PatternFill(fill_type='solid', fgColor=hex_color.lstrip('#').upper())

def _xl_side(color='E4E4E0'):
    from openpyxl.styles import Side
    return Side(style='thin', color=color.lstrip('#').upper())

def _xl_border_all(color='E4E4E0'):
    from openpyxl.styles import Border
    s = _xl_side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def _xl_border_bottom(color='E4E4E0'):
    from openpyxl.styles import Border
    return Border(bottom=_xl_side(color))

def _xl_align(h='left', v='center', wrap=False):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _xl_col_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def _xl_header_band(ws, title, meta_lines, n_cols):
    """Fila 1 oscura (título), fila 2 verde (metadatos)."""
    from openpyxl.styles import Font, Alignment
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 18
    for c in range(1, n_cols + 1):
        ws.cell(row=1, column=c).fill = _xl_fill(_C['dark'])
        ws.cell(row=2, column=c).fill = _xl_fill(_C['green_dark'])
    t = ws.cell(row=1, column=1, value=title)
    t.font  = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    m = ws.cell(row=2, column=1, value=meta_lines)
    m.font  = Font(name='Calibri', size=8, color='FFFFFF')
    m.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

def _xl_table_header(ws, row_num, headers, bg=None, wrap=False):
    ws.row_dimensions[row_num].height = 28 if wrap else 20
    bg = (bg or _C['dark']).lstrip('#').upper()
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=col, value=h)
        c.font      = _xl_font(bold=True, size=8, color='FFFFFF')
        c.fill      = _xl_fill(bg)
        c.alignment = _xl_align('center', 'center', wrap=wrap)
        c.border    = _xl_border_all(bg)

def _xl_data_row(ws, row_num, values, alt=False):
    ws.row_dimensions[row_num].height = 16
    bg = _C['bg'] if alt else _C['white']
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=v)
        c.font      = _xl_font(size=9)
        c.fill      = _xl_fill(bg)
        c.alignment = _xl_align('left', 'center')
        c.border    = _xl_border_bottom()

def _xl_status_cell(cell, estado):
    map_ = {
        'confirmado': (_C['green_light'],  _C['green_dark']),
        'pendiente':  (_C['orange_light'], _C['orange']),
        'rechazado':  (_C['red_light'],    _C['red']),
    }
    bg, fg = map_.get(estado, (_C['bg2'], _C['gray']))
    estado_label = {'confirmado': 'Confirmado', 'pendiente': 'Pendiente',
                    'rechazado': 'Rechazado'}.get(estado, estado)
    cell.value     = estado_label
    cell.fill      = _xl_fill(bg)
    cell.font      = _xl_font(bold=True, size=8, color=fg)
    cell.alignment = _xl_align('center', 'center')

def _xl_rank_row(ws, row_num, values, pos):
    """Fila de datos con resaltado dorado/plata/bronce para top 3."""
    colors_map = {
        1: (_C['gold_light'],   _C['gold'],   '1.'),
        2: (_C['silver_light'], _C['silver'], '2.'),
        3: (_C['bronze_light'], _C['bronze'], '3.'),
    }
    if pos in colors_map:
        bg, fg, _ = colors_map[pos]
        ws.row_dimensions[row_num].height = 18
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col, value=v)
            c.fill      = _xl_fill(bg)
            c.font      = _xl_font(bold=True, size=9, color=fg)
            c.alignment = _xl_align('center' if col == 1 else 'left', 'center')
            c.border    = _xl_border_bottom(fg)
    else:
        _xl_data_row(ws, row_num, values, alt=(pos % 2 == 0))


# ─── EXCEL: Reporte general ───────────────────────────────────────────────────

@admin_bp.route('/reportes/<id_quiniela>/csv', methods=['GET'])
@admin_required
def reporte_csv(id_quiniela):
    import openpyxl
    q            = Quiniela.query.get_or_404(id_quiniela)
    liga         = Liga.query.get(q.id_liga)
    partidos     = Partido.query.filter_by(id_quiniela=id_quiniela).order_by(Partido.inicio.asc()).all()
    participantes = UsuarioQuiniela.query.filter_by(id_quiniela=id_quiniela).order_by(
        UsuarioQuiniela.puntos_total.desc()).all()

    wb = openpyxl.Workbook()

    # ── Hoja 1: Predicciones ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Predicciones"
    ws.sheet_view.showGridLines = False

    n_cols  = 5 + len(partidos)
    meta    = (f"{liga.nombre if liga else 'Liga MX'}   |   Estado: {q.estado.capitalize()}   |   "
               f"Pozo: ${float(q.pozo_acumulado):,.0f}   |   "
               f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}")
    _xl_header_band(ws, f"Reporte de Predicciones  —  {q.nombre}", meta, n_cols)
    ws.row_dimensions[3].height = 8   # separador

    # Cabecera de tabla: partidos con nombre corto y fecha en wrap
    match_hdrs = []
    for i, p in enumerate(partidos, 1):
        el = Equipo.query.get(p.local)
        ev = Equipo.query.get(p.visitante)
        nl = el.nombre[:6] if el else '?'
        nv = ev.nombre[:6] if ev else '?'
        match_hdrs.append(f'P{i}\n{nl} v {nv}')
    _xl_table_header(ws, 4, ['#', 'Jugador', 'Correo', 'Pts', 'Pago'] + match_hdrs, wrap=True)

    for idx, uq in enumerate(participantes, 1):
        usr   = Usuario.query.get(uq.id_usr)
        if not usr: continue
        pago  = Pago.query.filter_by(id_usr=uq.id_usr, id_quiniela=id_quiniela).first()
        picks = []
        for p in partidos:
            pr = Prediccion.query.filter_by(id_usr=uq.id_usr, id_partido=p.id_partido).first()
            picks.append('/'.join(pr.selecciones) if pr else '-')
        rn = 4 + idx
        _xl_data_row(ws, rn, [idx, usr.nombre_completo, usr.correo, uq.puntos_total, ''] + picks, alt=(idx % 2 == 0))
        _xl_status_cell(ws.cell(row=rn, column=5), pago.estado if pago else 'sin_pago')
        ws.cell(row=rn, column=1).alignment = _xl_align('center', 'center')
        ws.cell(row=rn, column=4).alignment = _xl_align('center', 'center')

    ws.freeze_panes = 'A5'
    # Anchos: fijo para columnas de nombre/correo, pequeño para partidos
    pw = max(6, min(10, 90 // max(len(partidos), 1)))
    _xl_col_widths(ws, [4, 24, 30, 5, 13] + [pw] * len(partidos))

    # ── Hoja 2: Marcadores ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Marcadores")
    ws2.sheet_view.showGridLines = False
    meta2 = f"{q.nombre}   |   {liga.nombre if liga else ''}   |   {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    _xl_header_band(ws2, "Marcadores de Partidos", meta2, 6)
    ws2.row_dimensions[3].height = 8
    _xl_table_header(ws2, 4, ['#', 'Equipo Local', 'Resultado', 'Equipo Visitante', 'Fecha', 'Estado'],
                     bg=_C['green_dark'])

    for i, p in enumerate(partidos, 1):
        el = Equipo.query.get(p.local)
        ev = Equipo.query.get(p.visitante)
        if p.cancelado:
            score, status = 'Cancelado', 'Cancelado'
        elif p.ptos_local is not None:
            score  = f"{p.ptos_local}  -  {p.ptos_visitante}"
            status = 'Finalizado'
        else:
            score, status = '-', 'Pendiente'
        rn = 4 + i
        _xl_data_row(ws2, rn,
                     [i, el.nombre if el else '?', score, ev.nombre if ev else '?',
                      p.inicio.strftime('%d/%m/%Y  %H:%M'), status],
                     alt=(i % 2 == 0))
        sc = ws2.cell(row=rn, column=3)
        sc.alignment = _xl_align('center', 'center')
        if p.ptos_local is not None and not p.cancelado:
            sc.fill = _xl_fill(_C['green_light'])
            sc.font = _xl_font(bold=True, size=9, color=_C['green_dark'])
        st = ws2.cell(row=rn, column=6)
        st.alignment = _xl_align('center', 'center')
        if status == 'Finalizado':
            st.fill = _xl_fill(_C['green_light'])
            st.font = _xl_font(bold=True, size=8, color=_C['green_dark'])
        elif status == 'Pendiente':
            st.fill = _xl_fill(_C['orange_light'])
            st.font = _xl_font(bold=True, size=8, color=_C['orange'])
        elif status == 'Cancelado':
            st.fill = _xl_fill(_C['red_light'])
            st.font = _xl_font(bold=True, size=8, color=_C['red'])

    ws2.freeze_panes = 'A5'
    _xl_col_widths(ws2, [4, 24, 14, 24, 18, 13])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = q.nombre.replace(' ', '_').replace('—', '-')
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}_reporte.xlsx"'})


# ─── EXCEL: Pagos ─────────────────────────────────────────────────────────────

@admin_bp.route('/reportes/<id_quiniela>/pagos/csv', methods=['GET'])
@admin_required
def reporte_pagos_csv(id_quiniela):
    import openpyxl
    q    = Quiniela.query.get_or_404(id_quiniela)
    liga = Liga.query.get(q.id_liga)
    pagos = Pago.query.filter_by(id_quiniela=id_quiniela).all()

    confirmados     = sum(1 for p in pagos if p.estado == 'confirmado')
    total_recaudado = sum(p.monto for p in pagos if p.estado == 'confirmado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"
    ws.sheet_view.showGridLines = False

    meta = (f"{liga.nombre if liga else ''}   |   {confirmados}/{len(pagos)} confirmados   |   "
            f"Recaudado: ${float(total_recaudado):,.0f}   |   "
            f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}")
    _xl_header_band(ws, f"Reporte de Pagos  —  {q.nombre}", meta, 8)
    ws.row_dimensions[3].height = 8
    _xl_table_header(ws, 4, ['#', 'Jugador', 'Correo', 'Monto', 'Metodo', 'Estado', 'Fecha Pago', 'Confirmado'])

    for idx, pg in enumerate(pagos, 1):
        usr = Usuario.query.get(pg.id_usr)
        rn  = 4 + idx
        _xl_data_row(ws, rn, [
            idx,
            usr.nombre_completo if usr else '—',
            usr.correo if usr else '—',
            float(pg.monto),
            pg.metodo or '—',
            '',
            pg.fecha_pago.strftime('%d/%m/%Y  %H:%M') if pg.fecha_pago else '—',
            pg.fecha_confirmacion.strftime('%d/%m/%Y  %H:%M') if pg.fecha_confirmacion else '—',
        ], alt=(idx % 2 == 0))
        ws.cell(row=rn, column=1).alignment = _xl_align('center', 'center')
        mc = ws.cell(row=rn, column=4)
        mc.number_format = '"$"#,##0.00'
        mc.alignment     = _xl_align('right', 'center')
        mc.font          = _xl_font(bold=True, size=9)
        _xl_status_cell(ws.cell(row=rn, column=6), pg.estado)

    ws.freeze_panes = 'A5'
    _xl_col_widths(ws, [4, 26, 32, 12, 14, 14, 22, 22])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = q.nombre.replace(' ', '_').replace('—', '-')
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}_pagos.xlsx"'})


# ─── EXCEL: Posiciones ────────────────────────────────────────────────────────

@admin_bp.route('/reportes/<id_quiniela>/posiciones/csv', methods=['GET'])
@admin_required
def reporte_posiciones_csv(id_quiniela):
    import openpyxl
    q             = Quiniela.query.get_or_404(id_quiniela)
    liga          = Liga.query.get(q.id_liga)
    participantes = UsuarioQuiniela.query.filter_by(id_quiniela=id_quiniela).order_by(
        UsuarioQuiniela.puntos_total.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Posiciones"
    ws.sheet_view.showGridLines = False

    meta = (f"{liga.nombre if liga else ''}   |   Pozo: ${float(q.pozo_acumulado):,.0f}   |   "
            f"Estado: {q.estado.capitalize()}   |   "
            f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}")
    _xl_header_band(ws, f"Tabla de Posiciones  —  {q.nombre}", meta, 5)
    ws.row_dimensions[3].height = 8
    _xl_table_header(ws, 4, ['Pos.', 'Jugador', 'Correo', 'Puntos', 'Estado Pago'], bg=_C['green_dark'])

    for idx, uq in enumerate(participantes, 1):
        usr  = Usuario.query.get(uq.id_usr)
        if not usr: continue
        pago = Pago.query.filter_by(id_usr=uq.id_usr, id_quiniela=id_quiniela).first()
        estado = pago.estado if pago else 'sin_pago'
        rn = 4 + idx
        _xl_rank_row(ws, rn, [idx, usr.nombre_completo, usr.correo, uq.puntos_total, ''], pos=idx)
        _xl_status_cell(ws.cell(row=rn, column=5), estado)
        ws.cell(row=rn, column=1).alignment = _xl_align('center', 'center')
        ws.cell(row=rn, column=4).alignment = _xl_align('center', 'center')
        ws.cell(row=rn, column=4).font      = _xl_font(bold=True, size=10)

    ws.freeze_panes = 'A5'
    _xl_col_widths(ws, [6, 28, 34, 9, 16])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = q.nombre.replace(' ', '_').replace('—', '-')
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{fname}_posiciones.xlsx"'})


# ─── PDF HELPERS ──────────────────────────────────────────────────────────────

def _draw_page(canvas, doc, title, subtitle):
    """Header y footer en cada página del PDF."""
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    w, h = doc.pagesize
    # Banda oscura superior
    canvas.setFillColor(colors.HexColor(_C['dark']))
    canvas.rect(0, h - 2.0*cm, w, 2.0*cm, fill=1, stroke=0)
    # Línea acento verde
    canvas.setFillColor(colors.HexColor(_C['green']))
    canvas.rect(0, h - 2.15*cm, w, 0.15*cm, fill=1, stroke=0)
    # Barra lateral izquierda verde
    canvas.setFillColor(colors.HexColor(_C['green']))
    canvas.rect(0, h - 2.0*cm, 0.35*cm, 2.0*cm, fill=1, stroke=0)
    # Título
    canvas.setFillColor(colors.white)
    canvas.setFont('Helvetica-Bold', 12)
    canvas.drawString(0.7*cm, h - 1.35*cm, title)
    # Subtítulo
    canvas.setFont('Helvetica', 7.5)
    canvas.setFillColor(colors.HexColor('#aaaaaa'))
    canvas.drawString(0.7*cm, h - 1.80*cm, subtitle)
    # Footer
    canvas.setFillColor(colors.HexColor(_C['border']))
    canvas.rect(0, 0.6*cm, w, 0.04*cm, fill=1, stroke=0)
    canvas.setFont('Helvetica', 6.5)
    canvas.setFillColor(colors.HexColor(_C['gray']))
    canvas.drawString(0.7*cm, 0.3*cm, 'MABQUI Quinielas')
    canvas.drawRightString(w - 0.7*cm, 0.3*cm,
                           f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y a las %H:%M')}")


def _pdf_styles():
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Seccion', parent=styles['Normal'],
        fontSize=8, fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=3,
        textColor=colors.HexColor(_C['gray']),
        letterSpacing=0.8,
    ))
    return styles


def _pdf_info_strip(content_w, kv_pairs):
    """Tira horizontal con pares clave-valor (etiqueta gris + valor negro)."""
    from reportlab.platypus import Table as RLTable, TableStyle
    from reportlab.lib import colors
    n = len(kv_pairs)
    cw = content_w / n
    labels = [k for k, _ in kv_pairs]
    values = [v for _, v in kv_pairs]
    tbl = RLTable([labels, values], colWidths=[cw]*n, rowHeights=[11, 18])
    tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,-1), colors.HexColor(_C['bg2'])),
        ('TOPPADDING',   (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',(0,0), (-1,-1), 5),
        ('LEFTPADDING',  (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ('FONTNAME',     (0,0), (-1,0),  'Helvetica'),
        ('FONTSIZE',     (0,0), (-1,0),  6.5),
        ('TEXTCOLOR',    (0,0), (-1,0),  colors.HexColor(_C['gray'])),
        ('FONTNAME',     (0,1), (-1,1),  'Helvetica-Bold'),
        ('FONTSIZE',     (0,1), (-1,1),  11),
        ('TEXTCOLOR',    (0,1), (-1,1),  colors.HexColor(_C['dark'])),
        ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('LINEAFTER',    (0,0), (-2,-1), 0.5, colors.HexColor(_C['border'])),
        ('BOX',          (0,0), (-1,-1), 0.5, colors.HexColor(_C['border'])),
    ]))
    return tbl


def _pdf_table(data, col_widths, hdr_bg=None, status_col=None, rank_rows=False):
    """Tabla limpia: header de color, filas alternas, coloreado de estado."""
    from reportlab.platypus import Table as RLTable, TableStyle
    from reportlab.lib import colors

    hc    = colors.HexColor(hdr_bg or _C['dark'])
    white = colors.white
    alt   = colors.HexColor(_C['bg'])
    brd   = colors.HexColor(_C['border'])

    cmds = [
        # Header
        ('BACKGROUND',    (0,0), (-1,0),  hc),
        ('TEXTCOLOR',     (0,0), (-1,0),  white),
        ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0),  7.5),
        ('ALIGN',         (0,0), (-1,0),  'CENTER'),
        ('VALIGN',        (0,0), (-1,0),  'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,0),  5),
        ('BOTTOMPADDING', (0,0), (-1,0),  5),
        # Datos
        ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',      (0,1), (-1,-1), 8),
        ('TOPPADDING',    (0,1), (-1,-1), 4),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
        ('VALIGN',        (0,1), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [white, alt]),
        ('LINEBELOW',     (0,0), (-1,-1), 0.3, brd),
    ]

    # Top 3 con color
    if rank_rows:
        rank_colors = [
            (1, colors.HexColor(_C['gold_light']),   colors.HexColor(_C['gold'])),
            (2, colors.HexColor(_C['silver_light']), colors.HexColor(_C['silver'])),
            (3, colors.HexColor(_C['bronze_light']), colors.HexColor(_C['bronze'])),
        ]
        for pos, bg, fg in rank_colors:
            if pos < len(data):
                cmds += [
                    ('BACKGROUND', (0, pos), (-1, pos), bg),
                    ('FONTNAME',   (0, pos), (-1, pos), 'Helvetica-Bold'),
                    ('TEXTCOLOR',  (0, pos), (0, pos),  fg),
                ]

    # Columna de estado con colores de badge
    if status_col:
        ci = status_col - 1
        for ri, row in enumerate(data[1:], 1):
            if ci < len(row):
                val = str(row[ci]).lower()
                if val == 'confirmado':
                    cmds += [('BACKGROUND', (ci, ri), (ci, ri), colors.HexColor(_C['green_light'])),
                              ('TEXTCOLOR',  (ci, ri), (ci, ri), colors.HexColor(_C['green_dark'])),
                              ('FONTNAME',   (ci, ri), (ci, ri), 'Helvetica-Bold')]
                elif val == 'pendiente':
                    cmds += [('BACKGROUND', (ci, ri), (ci, ri), colors.HexColor(_C['orange_light'])),
                              ('TEXTCOLOR',  (ci, ri), (ci, ri), colors.HexColor(_C['orange'])),
                              ('FONTNAME',   (ci, ri), (ci, ri), 'Helvetica-Bold')]
                elif val == 'rechazado':
                    cmds += [('BACKGROUND', (ci, ri), (ci, ri), colors.HexColor(_C['red_light'])),
                              ('TEXTCOLOR',  (ci, ri), (ci, ri), colors.HexColor(_C['red'])),
                              ('FONTNAME',   (ci, ri), (ci, ri), 'Helvetica-Bold')]

    tbl = RLTable(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(cmds))
    return tbl


# ─── PDF: Reporte general ─────────────────────────────────────────────────────

@admin_bp.route('/reportes/<id_quiniela>/pdf', methods=['GET'])
@admin_required
def reporte_pdf(id_quiniela):
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    import functools

    q             = Quiniela.query.get_or_404(id_quiniela)
    liga          = Liga.query.get(q.id_liga)
    partidos      = Partido.query.filter_by(id_quiniela=id_quiniela).order_by(Partido.inicio.asc()).all()
    participantes = UsuarioQuiniela.query.filter_by(id_quiniela=id_quiniela).order_by(
        UsuarioQuiniela.puntos_total.desc()).all()

    PAGE    = landscape(A4)
    MARGIN  = 1.6*cm
    TOP_MAR = 2.6*cm
    BOT_MAR = 1.2*cm
    cw      = PAGE[0] - 2*MARGIN   # ancho útil

    title_str = q.nombre
    sub_str   = (f"{liga.nombre if liga else 'Liga MX'}   |   "
                 f"Estado: {q.estado.capitalize()}   |   "
                 f"Pozo: ${float(q.pozo_acumulado):,.0f}")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            leftMargin=MARGIN, rightMargin=MARGIN,
                            topMargin=TOP_MAR, bottomMargin=BOT_MAR)
    fn  = functools.partial(_draw_page, title=title_str, subtitle=sub_str)
    styles = _pdf_styles()
    story  = []

    # Tira de info
    confirmados = sum(1 for uq in participantes
                      if Pago.query.filter_by(id_usr=uq.id_usr, id_quiniela=id_quiniela, estado='confirmado').first())
    story.append(Spacer(1, 0.25*cm))
    story.append(_pdf_info_strip(cw, [
        ('JUGADORES',    str(len(participantes))),
        ('PARTIDOS',     str(len(partidos))),
        ('POZO TOTAL',   f"${float(q.pozo_acumulado):,.0f}"),
        ('CONFIRMADOS',  str(confirmados)),
        ('ESTADO',       q.estado.capitalize()),
        ('CIERRE',       q.cierre.strftime('%d/%m/%Y %H:%M') if q.cierre else '—'),
    ]))
    story.append(Spacer(1, 0.4*cm))

    # Sección predicciones
    story.append(Paragraph("PREDICCIONES POR JUGADOR", styles['SeccionTitle'] if 'SeccionTitle' in styles else styles['SeccionTitle'] if False else styles['Seccion']))
    story.append(HRFlowable(width=cw, thickness=1.5, color=colors.HexColor(_C['green']), spaceAfter=5))

    # Calcular anchos: columnas fijas + columnas de partido iguales
    n_p      = len(partidos)
    fixed_w  = 0.7*cm + 4.2*cm + 1.0*cm + 1.6*cm   # #, jugador, pts, pago
    part_avail = cw - fixed_w
    pw       = min(part_avail / max(n_p, 1), 1.8*cm)

    hdr_p = ['#', 'Jugador', 'Pts', 'Pago']
    for i, p in enumerate(partidos, 1):
        el = Equipo.query.get(p.local)
        ev = Equipo.query.get(p.visitante)
        hdr_p.append(f'P{i}')   # número del partido; leyenda abajo

    rows_p = [hdr_p]
    for idx, uq in enumerate(participantes, 1):
        usr  = Usuario.query.get(uq.id_usr)
        if not usr: continue
        pago = Pago.query.filter_by(id_usr=uq.id_usr, id_quiniela=id_quiniela).first()
        row  = [str(idx), usr.nombre_completo[:26], str(uq.puntos_total),
                pago.estado if pago else 'sin pago']
        for p in partidos:
            pr = Prediccion.query.filter_by(id_usr=uq.id_usr, id_partido=p.id_partido).first()
            row.append('/'.join(pr.selecciones) if pr else '-')
        rows_p.append(row)

    story.append(_pdf_table(rows_p,
                             [0.7*cm, 4.2*cm, 1.0*cm, 1.6*cm] + [pw]*n_p,
                             status_col=4, rank_rows=True))
    story.append(Spacer(1, 0.5*cm))

    # Leyenda de partidos
    story.append(Paragraph("PARTIDOS  (L = Local gana  |  E = Empate  |  V = Visitante gana)", styles['Seccion']))
    story.append(HRFlowable(width=cw, thickness=1.5, color=colors.HexColor(_C['green']), spaceAfter=5))

    leg_rows = [['No.', 'Local', 'Visitante', 'Fecha', 'Resultado']]
    for i, p in enumerate(partidos, 1):
        el = Equipo.query.get(p.local)
        ev = Equipo.query.get(p.visitante)
        if p.cancelado:
            res = 'Cancelado'
        elif p.ptos_local is not None:
            res = f"{p.ptos_local} - {p.ptos_visitante}"
        else:
            res = 'Pendiente'
        leg_rows.append([f'P{i}', el.nombre if el else '?', ev.nombre if ev else '?',
                          p.inicio.strftime('%d/%m  %H:%M'), res])

    story.append(_pdf_table(leg_rows,
                             [0.8*cm, 5.0*cm, 5.0*cm, 2.2*cm, 2.4*cm],
                             hdr_bg=_C['green_dark']))

    doc.build(story, onFirstPage=fn, onLaterPages=fn)
    buf.seek(0)
    fname = q.nombre.replace(' ', '_').replace('—', '-')
    return Response(buf.getvalue(), mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{fname}_reporte.pdf"'})


# ─── PDF: Posiciones ──────────────────────────────────────────────────────────

@admin_bp.route('/reportes/<id_quiniela>/posiciones/pdf', methods=['GET'])
@admin_required
def reporte_posiciones_pdf(id_quiniela):
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    import functools

    q             = Quiniela.query.get_or_404(id_quiniela)
    liga          = Liga.query.get(q.id_liga)
    participantes = UsuarioQuiniela.query.filter_by(id_quiniela=id_quiniela).order_by(
        UsuarioQuiniela.puntos_total.desc()).all()

    PAGE    = A4
    MARGIN  = 1.8*cm
    cw      = PAGE[0] - 2*MARGIN

    lider     = participantes[0] if participantes else None
    lider_usr = Usuario.query.get(lider.id_usr) if lider else None

    title_str = f"Tabla de Posiciones  —  {q.nombre}"
    sub_str   = (f"{liga.nombre if liga else 'Liga MX'}   |   "
                 f"Pozo: ${float(q.pozo_acumulado):,.0f}   |   "
                 f"Estado: {q.estado.capitalize()}")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            leftMargin=MARGIN, rightMargin=MARGIN,
                            topMargin=2.6*cm, bottomMargin=1.2*cm)
    fn  = functools.partial(_draw_page, title=title_str, subtitle=sub_str)
    styles = _pdf_styles()
    story  = []

    story.append(Spacer(1, 0.25*cm))
    story.append(_pdf_info_strip(cw, [
        ('JUGADORES',    str(len(participantes))),
        ('POZO',         f"${float(q.pozo_acumulado):,.0f}"),
        ('LIDER',        lider_usr.nombre_completo.split()[0] if lider_usr else '—'),
        ('PUNTOS LIDER', str(lider.puntos_total) if lider else '—'),
        ('ESTADO',       q.estado.capitalize()),
    ]))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("CLASIFICACION GENERAL", styles['Seccion']))
    story.append(HRFlowable(width=cw, thickness=1.5, color=colors.HexColor(_C['green']), spaceAfter=5))

    rows = [['Pos.', 'Jugador', 'Correo', 'Puntos', 'Estado Pago']]
    for idx, uq in enumerate(participantes, 1):
        usr  = Usuario.query.get(uq.id_usr)
        if not usr: continue
        pago = Pago.query.filter_by(id_usr=uq.id_usr, id_quiniela=id_quiniela).first()
        rows.append([
            f'{idx}.',
            usr.nombre_completo,
            usr.correo,
            str(uq.puntos_total),
            pago.estado if pago else 'sin pago',
        ])

    story.append(_pdf_table(rows,
                             [1.2*cm, 5.0*cm, 5.8*cm, 1.5*cm, 2.5*cm],
                             hdr_bg=_C['green_dark'],
                             rank_rows=True,
                             status_col=5))

    doc.build(story, onFirstPage=fn, onLaterPages=fn)
    buf.seek(0)
    fname = q.nombre.replace(' ', '_').replace('—', '-')
    return Response(buf.getvalue(), mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="{fname}_posiciones.pdf"'})
